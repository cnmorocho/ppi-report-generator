import os
from openpyxl import Workbook, load_workbook

class ReportGenerator:
    def __init__(self, absolute_path):
        if not absolute_path.endswith(".xlsx"):
            raise ValueError("Invalid file extension. Please provide a valid excel")
        if os.path.exists(absolute_path):
            self.workbook = load_workbook(absolute_path)
        else:
            self.workbook = Workbook(absolute_path)
        self.worksheet = self.workbook.active
        self.absolute_path = absolute_path

